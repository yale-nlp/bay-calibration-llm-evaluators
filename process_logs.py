import os
from pathlib import Path
import math
import openpyxl

log_dir = Path('./results/logs')
log_paths = list(log_dir.glob('*-27/*.log'))

def process_logs(log_paths):
    our_diff_w_p = []
    their_diff_w_p = []
    for log_path in log_paths:
        with open(log_path) as log_file:
            for line in log_file:
                if 'and estimated p mean:' in line:
                    our_diff_w_p.append(float(line.split()[-1]))
                elif 'Difference between true p and k:' in line:
                    their_diff_w_p.append(float(line.split()[-1]))
        print(f'{log_path} done')

        our_not_nan, our_better = [], 0
        our_nans = 0
        for x, y in zip(our_diff_w_p, their_diff_w_p):
            if not math.isnan(x) and x < y:
                our_better += 1
            if not math.isnan(x):
                our_not_nan.append(x)
            else:
                our_nans += 1
        print(f'Our mean: {sum(our_not_nan) / len(our_not_nan)}')
        print(f'Their mean: {sum(their_diff_w_p) / len(their_diff_w_p)}')
        print(f'Our nans: {our_nans} / {len(our_diff_w_p)} logs ({our_nans / len(our_diff_w_p) * 100:.2f}%)')
        print(f'Our is better for { our_better } / {len(our_diff_w_p)} logs ({our_better / len(our_diff_w_p) * 100:.2f}%)')

import re
import pandas as pd

# Function to read the document from a file and extract data
def extract_data_from_file(file_path, human=False):
    with open(file_path, 'r') as file:
        document = file.read()

    sections = document.split("**************************")[1:]  # Skip the first empty split
    data = []

    if 'cv' in file_path.stem.lower():
        # get only cv average results
        for i in range(len(sections)):
            if 'CV average:' in sections[i]:
                sections[i] = sections[i].split('CV average:')[1]

    for section in sections:
        models = re.search(r"Comparing\s+([^\s]+)\s+and\s+([^\s]+)", section)
        if models:
            model1, model2 = models.groups()
            model1, model2 = model1.rstrip('.').replace('..', ''), model2.rstrip('.').replace('..', '')

        mean_diff_search = re.search(r"Difference between true p and estimated p mean: ([\d.]+|nan)", section)
        k_diff_search = re.search(r"Difference between true p and k: ([\d.]+|nan)", section)
        human_label_p_diff_search = re.search(r"phat by human label error: ([\d.]+)", section)

        if mean_diff_search and k_diff_search:
            mean_diff = mean_diff_search.group(1)
            k_diff = k_diff_search.group(1)
            human_label_p_diff = human_label_p_diff_search.group(1) if human_label_p_diff_search else None

            # Check if 'nan' is present and handle it
            mean_diff = None if mean_diff.lower() == 'nan' else float(mean_diff)
            k_diff = None if k_diff.lower() == 'nan' else float(k_diff)
            if human_label_p_diff is not None:
                human_label_p_diff = None if human_label_p_diff.lower() == 'nan' else float(human_label_p_diff)

            if human:
                data.append({
                    "Model 1": model1.strip(),
                    "Model 2": model2.strip(),
                    "Difference Mean": mean_diff,
                    "Difference K": k_diff,
                    "phat by human label": human_label_p_diff,
                })
            else:
                data.append({
                    "Model 1": model1.strip(),
                    "Model 2": model2.strip(),
                    "Difference Mean": mean_diff,
                    "Difference K": k_diff,
                })

    df = pd.DataFrame(data)
    if len(df) == 0:
        return
    if human:
        df = pd.concat([df, pd.DataFrame([{'Model 1': 'mean', 'Model 2': 'mean', 'Difference Mean': df['Difference Mean'].mean(), 'Difference K': df['Difference K'].mean(), 'phat by human label': df['phat by human label'].mean()}])])
    else:
        df = pd.concat([df, pd.DataFrame([{'Model 1': 'mean', 'Model 2': 'mean', 'Difference Mean': df['Difference Mean'].mean(), 'Difference K': df['Difference K'].mean()}])])
    return df
    

def merge_methods(dfs, human=False):
    all_dfs = []
    all_idx_to_color = []
    for dataset_name, dfs_list in dfs.items():
        merged_dfs = []
        for method_name, settings_name, df in dfs_list:
            # skipping Bayesian in-dist
            if method_name == 'bayds' and 'in_dist' in settings_name:
                continue
            df.rename(columns={'Difference Mean': f'{method_name} {settings_name} |Mean-p|', 'Difference K': f'{method_name} {settings_name} |k-p|', 'phat by human label': f'{method_name} {settings_name} |human-p|'}, inplace=True)
            merged_dfs.append(df)
        merged_df = pd.concat(merged_dfs)
        # Nan values are not considered in the mean calculation
        merged_df = merged_df.groupby(['Model 1', 'Model 2']).mean().reset_index()
        # arrange model1 and model2 in alphabetical order
        for i in range(len(merged_df)):
            if merged_df.loc[i, 'Model 1'] > merged_df.loc[i, 'Model 2']:
                merged_df.loc[i, 'Model 1'], merged_df.loc[i, 'Model 2'] = merged_df.loc[i, 'Model 2'], merged_df.loc[i, 'Model 1']
        merged_df = merged_df.sort_values(by=['Model 1', 'Model 2'])
        # Record idx of the better value in each setting
        merged_df.sort_index(axis=1, inplace=True)
        row_idx = merged_df.loc[merged_df['Model 1'] == 'mean'].index.item()
        col_idx = []
        step_size = 3 if human else 2
        ## compare each 3 columns and record the best one
        for i in range(2, len(merged_df.columns), step_size):
            best_col = merged_df.columns[i]
            best_val = merged_df.loc[row_idx, best_col]
            for j in range(0, step_size):
                if i + j >= len(merged_df.columns):
                    break
                col = merged_df.columns[i + j]
                val = merged_df.loc[row_idx, col]
                if val < best_val:
                    best_col = col
                    best_val = val
            col_idx.append((best_col, merged_df.columns.get_loc(best_col)))
            
        os.makedirs('./results/csv', exist_ok=True)
        merged_df.sort_index(axis=1).to_csv(f'./results/csv/{dataset_name}-merged.csv', index=False)
        all_dfs.append((dataset_name, merged_df))
        all_idx_to_color.append((row_idx, col_idx))
    with pd.ExcelWriter('./results/csv/results-merged.xlsx') as writer:
        for dataset_name, merged_df in all_dfs:
            merged_df.sort_index(axis=1).to_excel(writer, sheet_name=dataset_name)
    # Color the recorded positions
    wb = openpyxl.load_workbook('./results/csv/results-merged.xlsx')
    for i, sheet in enumerate(wb.sheetnames):
        ws = wb[sheet]
        row_idx, col_idx = all_idx_to_color[i]
        for col_name, idx in col_idx:
            color = 'FF00FF00' if 'Mean' in col_name else 'FFFF0000'
            ws.cell(row=row_idx + 2, column=idx + 2).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
    wb.save('./results/csv/results-merged.xlsx')


def merge_methods_cv_only(dfs):
    all_dfs = []
    all_idx_to_color = []
    for dataset_name, dfs_list in dfs.items():
        merged_dfs = []
        for method_name, settings_name, df in dfs_list:
            # skipping Bayesian in-dist
            if method_name == 'bayds' and 'in_dist' in settings_name or 'cv' not in settings_name:
                continue
            df.rename(columns={'Difference Mean': f'{method_name} {settings_name} |Mean-p|', 'Difference K': f'{method_name} {settings_name} |k-p|'}, inplace=True)
            merged_dfs.append(df)
        merged_df = pd.concat(merged_dfs)
        # Nan values are not considered in the mean calculation
        merged_df = merged_df.groupby(['Model 1', 'Model 2']).mean().reset_index()
        # arrange model1 and model2 in alphabetical order
        for i in range(len(merged_df)):
            if merged_df.loc[i, 'Model 1'] > merged_df.loc[i, 'Model 2']:
                merged_df.loc[i, 'Model 1'], merged_df.loc[i, 'Model 2'] = merged_df.loc[i, 'Model 2'], merged_df.loc[i, 'Model 1']
        merged_df = merged_df.sort_values(by=['Model 1', 'Model 2'])
        # Record idx of the better value in each setting
        merged_df.sort_index(axis=1, inplace=True)
        row_idx = merged_df.loc[merged_df['Model 1'] == 'mean'].index.item()
        col_idx = []
        # compare each pair of columns and record the better one
        last_col, this_col = None, None
        for col in merged_df.columns:
            if '-p|' not in col:
                continue
            if last_col is None:
                last_col = col
                continue
            this_col = col
            if last_col.split()[0:2] == this_col.split()[0:2]:
                if merged_df.loc[row_idx, last_col] < merged_df.loc[row_idx, this_col]:
                    # last_col is better
                    col_idx.append((last_col, merged_df.columns.get_loc(last_col)))
                else:
                    # this_col is better
                    col_idx.append((this_col, merged_df.columns.get_loc(this_col)))
            last_col = this_col
        
        os.makedirs('./results/csv', exist_ok=True)
        merged_df.sort_index(axis=1).to_csv(f'./results/csv/{dataset_name}-merged.csv', index=False)
        all_dfs.append((dataset_name, merged_df))
        all_idx_to_color.append((row_idx, col_idx))
    with pd.ExcelWriter('./results/csv/results-merged.xlsx') as writer:
        for dataset_name, merged_df in all_dfs:
            merged_df.sort_index(axis=1).to_excel(writer, sheet_name=dataset_name)
    # Color the recorded positions
    wb = openpyxl.load_workbook('./results/csv/results-merged.xlsx')
    for i, sheet in enumerate(wb.sheetnames):
        ws = wb[sheet]
        row_idx, col_idx = all_idx_to_color[i]
        for col_name, idx in col_idx:
            color = 'FF00FF00' if 'Mean' in col_name else 'FFFF0000'
            ws.cell(row=row_idx + 2, column=idx + 2).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
    wb.save('./results/csv/results-merged-cvs.xlsx')

def main():
    # process every .log file under results/logs and save results into results/csv
    dfs = {}
    for log_path in log_paths:
        subdir = str(log_path.parent).split('results/logs/')[1] if len(str(log_path.parent).split('results/logs/')) > 1 else ''
        df = extract_data_from_file(log_path)
        if df is None:
            continue
        dataset_name, method_name = subdir.split('-')[0], subdir.split('-')[1]
        dfs.setdefault(dataset_name, []).append((method_name, log_path.stem, df))
        # df.to_csv(output_path, index=False)
        # print(f'{log_path} done')
    merge_methods(dfs)

if __name__ == '__main__':
    main()
